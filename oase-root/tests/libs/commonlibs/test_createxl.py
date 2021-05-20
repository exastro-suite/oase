
# Copyright 2019 NEC Corporation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

"""

createxl.pyのテスト

"""

import datetime
import os
import traceback
import unittest
import pytest
import pytz

from importlib import import_module
from unittest.mock import MagicMock, patch
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection

from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.test import override_settings

from django.urls import reverse
from libs.commonlibs.createxl import DecisionTableFactory
from web_app.models.models import DataObject, RuleType
from web_app.templatetags.common import get_message

class wsheetMock:
    def cell_type(self):
        return 1


def set_data_object():
    """
    テストデータ作成
    """
    now = datetime.datetime.now(pytz.timezone('UTC'))
    try:
        with transaction.atomic():
            # ルール種別
            RuleType(
                rule_type_id=1,
                rule_type_name='pytest',
                summary='',
                rule_table_name='pytesttable',
                generation_limit='1',
                group_id='1',
                artifact_id='1',
                container_id_prefix_staging='',
                container_id_prefix_product='',
                current_container_id_staging='',
                current_container_id_product='',
                label_count=1,
                disuse_flag='0',
                last_update_user='pytest',
                last_update_timestamp=now,
            ).save(force_insert=True)

            # データオブジェクト
            DataObject(
                data_object_id=1,
                rule_type_id=1,
                conditional_name='テスト条件名',
                label='label0',
                conditional_expression_id=1,
                last_update_user='pytest',
                last_update_timestamp=now,
            ).save(force_insert=True)

    except Exception as e:
        print(e)


def del_data_object():
    """
    テストデータ削除
    """
    RuleType.objects.all().delete()
    DataObject.objects.all().delete()

@pytest.mark.django_db
def set_DecisionTableFactory():
    """
    classの初期設定
    """
    dt_create = DecisionTableFactory(
            1,
            'testrule',
            'com.oase.pytesttable',
            'pytesttable',
            'pytesttableObject',
            'RuleTable pytesttable',
            '/root/pytest'
    )
    return dt_create


@pytest.mark.django_db
def test__get_action_count_info_true():
    """
    アクション実行／リトライの回数／間隔判定
    条件回数／条件期間(秒)
    大グループ優先順位／小グループ優先順位
    Trueになる場合
    """
    del_data_object()
    set_data_object()
    dt_create = set_DecisionTableFactory()

    strs = ["アクションリトライ間隔（必須）", "※「1」以上を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["アクションリトライ回数（必須）", "※「1」以上を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["アクション抑止間隔（必須）", "※「0」以上を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["アクション抑止回数（必須）", "※「0」以上を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["アクション条件回数（必須）", "※「1」以上を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["アクション条件期間(秒)（必須）", "※「1」以上を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["大グループ優先順位（必須）", "※「1」からの通番を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    strs = ["小グループ優先順位（必須）", "※「1」からの通番を定義して下さい。"]
    result = dt_create._get_action_count_info(strs)
    assert result == True

    del_data_object()

@pytest.mark.django_db
def test__get_action_count_info_false():
    """
    アクション実行／リトライの回数／間隔判定
    大グループ／小グループ
    Falseになる場合
    """
    del_data_object()
    set_data_object()
    dt_create = set_DecisionTableFactory()

    strs = ["アクション実行前パラメータ情報（必須）", "※ダブルクオーテーションは使用不可", "※不要の場合は「X」を定義"]
    result = dt_create._get_action_count_info(strs)
    assert result == False

    strs = ["有効日"]
    result = dt_create._get_action_count_info(strs)
    assert result == False

    strs = ["大グループ（必須）"]
    result = dt_create._get_action_count_info(strs)
    assert result == False


    strs = ["小グループ（必須）"]
    result = dt_create._get_action_count_info(strs)
    assert result == False

    del_data_object()


@pytest.mark.django_db
def test__set_width():
    """
    幅調整のテスト
    """

    del_data_object()
    set_data_object()

    dt_create = set_DecisionTableFactory()

    dt_create._create_condition()
    dt_create._create_header()
    dt_create._set_styles()
    ws = dt_create.tables_ws

    dt_create._set_width()

    for c in range(8, 12):
        if c != 12:
            assert ws.column_dimensions[get_column_letter(c)].width == 2.88
        else:
            assert ws.column_dimensions[get_column_letter(c)].width == None

    del_data_object()

@pytest.mark.django_db
def test__set_styles():
    """
    左揃えのテスト
    """

    del_data_object()
    set_data_object()

    dt_create = set_DecisionTableFactory()

    dt_create._create_condition()
    dt_create._create_header()

    ws = dt_create.tables_ws
    dt_create._set_styles()

    al_lb = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

    for c in range(2, 12):
        for r in range(12, 19):
            if c == 6 or c == 7:
                assert ws.cell(row=r, column=c).alignment == al_lb
            else:
                assert ws.cell(row=r, column=c).alignment != al_lb

    del_data_object()
