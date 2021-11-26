# Copyright 2019 NEC Corporation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# 
#     http://www.apache.org/licenses/LICENSE-2.0
# 
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#



"""
[概要]
   ディシジョンテーブルのエクセルファイルを作成する

"""


import os
import sys
import openpyxl
import pprint
import django
import traceback

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile

from django.conf import settings

from libs.commonlibs.define import *
from libs.commonlibs.aes_cipher import AESCipher
from libs.commonlibs.oase_logger import OaseLogger
from web_app.models.models import DataObject, RuleType, ConditionalExpression, System, ActionType, DriverType
from web_app.templatetags.common import get_message
from web_app.views.top.login import _get_system_lang_mode

logger = OaseLogger.get_instance()

class DecisionTableFactory:
    """
    [クラス概要]
    DecisionTableを作成するクラス
    rule_type_idからルール種別、データオブジェクトを取得し、
    template.xlsxファイルを基に作成する。
    アクション部はテンプレートをそのまま使えば良い。
    """

    # 多言語対応
    MSG_INFO = {
        'Tables'  : {
            'B11' : ['MOSJA11064',],
            'B19' : ['MOSJA11081',],
            #'B20' : ['MOSJA11082', {'rule_row_max':0}],
            'D11' : ['MOSJA11065',],
            'E11' : ['MOSJA11147',],
            'F11' : ['MOSJA11148',],
            'G11' : ['MOSJA11066',],
            'H11' : ['MOSJA11067',],
            'I11' : ['MOSJA11068',],
            'J11' : ['MOSJA11069',],
            'K11' : ['MOSJA11070',],
            'L11' : ['MOSJA11071',],
            'M11' : ['MOSJA11072',],
            'N11' : ['MOSJA11073',],
            'O11' : ['MOSJA11074',],
            'P11' : ['MOSJA11075',],
            'Q11' : ['MOSJA11076',],
            'R11' : ['MOSJA11077',],
            'S11' : ['MOSJA11078',],
            'T11' : ['MOSJA11079',],
            'U11' : ['MOSJA11080',],
        },
        'example' : {
            # 見出し行
            'B2'  : ['MOSJA11083',],
            'C2'  : ['MOSJA11084',],
            'D2'  : ['MOSJA11085',],
            'E2'  : ['MOSJA11147',],
            'F2'  : ['MOSJA11148',],
            'G2'  : ['MOSJA11086',],
            'H2'  : ['MOSJA11087',],
            'I2'  : ['MOSJA11088',],
            'J2'  : ['MOSJA11089',],
            'K2'  : ['MOSJA11090',],
            'L2'  : ['MOSJA11091',],
            'M2'  : ['MOSJA11092',],
            'N2'  : ['MOSJA11093',],
            'O2'  : ['MOSJA11094',],
            'P2'  : ['MOSJA11095',],
            'Q2'  : ['MOSJA11096',],
            'R2'  : ['MOSJA11097',],
            'S2'  : ['MOSJA11098',],
            'T2'  : ['MOSJA11099',],
            'U2'  : ['MOSJA11100',],
            # 具体例
            'C3'  : ['MOSJA11102',],
            'C4'  : ['MOSJA11102',],
            'C5'  : ['MOSJA11102',],
            'C6'  : ['MOSJA11102',],
            'C7'  : ['MOSJA11102',],
            'C8'  : ['MOSJA11102',],
            'C9'  : ['MOSJA11102',],
            'C10' : ['MOSJA11102',],
            'C11' : ['MOSJA11102',],
            'C12' : ['MOSJA11102',],
            'C13' : ['MOSJA11102',],
            'C14' : ['MOSJA11102',],
            'C15' : ['MOSJA11102',],
            'C16' : ['MOSJA11102',],
            'C17' : ['MOSJA11102',],
            'C18' : ['MOSJA11102',],
            'C19' : ['MOSJA11102',],
            'C20' : ['MOSJA11102',],
            'C21' : ['MOSJA11102',],
            'C22' : ['MOSJA11102',],
            'E3'  : ['MOSJA11151',],
            'E4'  : ['MOSJA11151',],
            'E5'  : ['MOSJA11151',],
            'E6'  : ['MOSJA11151',],
            'E7'  : ['MOSJA11151',],
            'E8'  : ['MOSJA11151',],
            'E9'  : ['MOSJA11151',],
            'E10' : ['MOSJA11151',],
            'E11' : ['MOSJA11151',],
            'E12' : ['MOSJA11151',],
            'E13' : ['MOSJA11151',],
            'E14' : ['MOSJA11151',],
            'E15' : ['MOSJA11151',],
            'E16' : ['MOSJA11151',],
            'E17' : ['MOSJA11151',],
            'E18' : ['MOSJA11151',],
            'E20' : ['MOSJA11151',],
            'E22' : ['MOSJA11151',],
            'F3'  : ['MOSJA11152',],
            'F4'  : ['MOSJA11152',],
            'F5'  : ['MOSJA11152',],
            'F6'  : ['MOSJA11152',],
            'F7'  : ['MOSJA11152',],
            'F8'  : ['MOSJA11152',],
            'F9'  : ['MOSJA11152',],
            'F10' : ['MOSJA11152',],
            'F11' : ['MOSJA11152',],
            'F12' : ['MOSJA11152',],
            'F13' : ['MOSJA11152',],
            'F14' : ['MOSJA11153',],
            'F15' : ['MOSJA11153',],
            'F16' : ['MOSJA11154',],
            'F17' : ['MOSJA11155',],
            'F18' : ['MOSJA11156',],
            'F20' : ['MOSJA11152',],
            'F22' : ['MOSJA11152',],
            'G19' : ['MOSJA11103',],
            'G21' : ['MOSJA11103',],
            # 記述例コメント
            'B25' : ['MOSJA11104',],
            'C25' : ['MOSJA11105',],
            'C26' : ['MOSJA11115',],
            'D25' : ['MOSJA11106',],
            'E25' : ['MOSJA11157',],
            'F25' : ['MOSJA11158',],
            'G25' : ['MOSJA11107',],
            'H25' : ['MOSJA11108',],
            'H26' : ['MOSJA11116',],
            'H27' : ['MOSJA11117',],
            'H28' : ['MOSJA11118',],
            'H29' : ['MOSJA11119',],
            'I25' : ['MOSJA11109',],
            'J25' : ['MOSJA11110',],
            'P25' : ['MOSJA11111',],
            'R25' : ['MOSJA11112',],
            'T25' : ['MOSJA11113',],
            'U25' : ['MOSJA11114',],
        },
    }

    def __init__(self, rule_type_id, rule_set, table_name, class_name, fact_name, save_path):
        """
        [メソッド概要]
        B2:E5 = ヘッダーのテンプレート
        C7:C18 = 条件部のテンプレート
        B21:L32 = アクション部のテンプレート
        """
        try:
            if rule_type_id and  not isinstance(rule_type_id, int):
                raise TypeError('rule_type_id:expected int, got %s' % type(rule_type_id).__name__)

            rule_type = RuleType.objects.get(rule_type_id=rule_type_id)
            self.data_object_list = DataObject.objects.filter(rule_type_id=rule_type_id).order_by('data_object_id')
            self.from_to = DataObject.objects.filter(rule_type_id=rule_type_id, conditional_expression_id=15)
        except Exception as e:
            logger.system_log('LOSM12067', rule_type_id, traceback.format_exc())

        # 初期化
        self.rule_type_id = rule_type_id
        self.rule_set = rule_set
        self.table_name = table_name
        self.class_name = class_name
        self.fact_name = fact_name
        self.save_path = save_path
        self.len_act = 18 # アクション部の数
        self.len_condition = len(self.data_object_list) + len(self.from_to) # 条件部の数
        self.lang = _get_system_lang_mode()

        # テンプレート読み込み
        cur_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
        self.wb = openpyxl.load_workbook(cur_dir + '/template.xlsx')
        self.tables_ws = self.wb['Tables']
        self.example_ws = self.wb['example']
        self.lists_ws = self.wb['Lists']
        self.condition = self.tables_ws['C7:C18']
        self.action = self.tables_ws['B22:L25']

        # 多言語対応
        for ws_name, v1 in self.MSG_INFO.items():
            for cell, msg in v1.items():
                # ルール上限数記述
                """
                if cell == 'B20':
                    rule_row_max = System.objects.get(config_id='RULE_ROW_MAX').value
                    msg[1]['rule_row_max'] = rule_row_max
                """

                if len(msg) >= 2:
                    self.wb[ws_name][cell].value = get_message(msg[0], self.lang, showMsgId=False, **(msg[1]))

                else:
                    self.wb[ws_name][cell].value = get_message(msg[0], self.lang, showMsgId=False)


    def _create_header(self):
        """
        [メソッド概要]
        ヘッダーを作成する
        """
        # ヘッダーにあるルールセットとSequentialを設定する
        self.tables_ws['D2'] = self.rule_set
        self.tables_ws['D5'] = 'true'

    def _create_condition(self):
        """
        [メソッド概要]
        条件部を作成する
        """
        # RuleTableを設定する
        self.tables_ws['C7'] = self.fact_name

        # 条件のカラムを追加する
        self.tables_ws.insert_cols(4, amount=self.len_condition - 1)

        #labelの数だけ条件部を作成する
        ce_list = ConditionalExpression.objects.all()
        cond_exp_dict = { 
            c.pk:[c.operator, get_message(c.description, self.lang, showMsgId=False)] 
            for c in ce_list 
        }

        diff = 3 
        for i, d in enumerate(self.data_object_list):
            ce_id = d.conditional_expression_id
            ce = cond_exp_dict[ce_id]
            if not ce_id == 15:
                formula = ce[0].format(d.label)
                self.tables_ws.cell(row = 8, column = i+diff, value = 'CONDITION')
                self.tables_ws.cell(row = 10, column = i+diff, value = formula)
                self.tables_ws.cell(row = 11, column = i+diff, value = d.conditional_name+ce[1])
            else:
                # 日時(From-To)の処理
                formula = ce[0].format(d.label, d.label)
                formulas = formula.split()
                self.tables_ws.cell(row = 8, column = i+diff, value = 'CONDITION')
                self.tables_ws.cell(row = 10, column = i+diff, value = formulas[0])
                self.tables_ws.cell(row = 11, column = i+diff, value = d.conditional_name+ce[1]+'(from)')
                diff += 1
                self.tables_ws.cell(row = 8, column = i+diff, value = 'CONDITION')
                self.tables_ws.cell(row = 10, column = i+diff, value = formulas[1])
                self.tables_ws.cell(row = 11, column = i+diff, value = d.conditional_name+ce[1]+'(to)')

        self.tables_ws['C7'] = 'RuleTable ' + self.table_name
        self.tables_ws['C9'] ='m:' + self.class_name

        #DataObject部(9行目)を結合
        self.tables_ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=2+self.len_condition)


    def _set_styles(self):
        """
        [メソッド概要]
        セルの塗りつぶし
        """
        #----------------
        # 初期設定
        #----------------
        # 色の設定
        fill_black = PatternFill(fill_type='solid', fgColor='00000000')
        fill_gray = PatternFill(fill_type='solid', fgColor='00bfbfbf')
        fill_orange = PatternFill(fill_type='solid', fgColor='ffcc99')
        fill_yellow = PatternFill(fill_type='solid', fgColor='ffff8f')
        fill_pink = PatternFill(fill_type='solid', fgColor='da9694')
        fill_lightblue = PatternFill(fill_type='solid', fgColor='ccffff')

        # 枠線の設定
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

        # fontの設定
        font = openpyxl.styles.Font(name='Arial', size=7)
        font_bold = openpyxl.styles.Font(name='Arial', size=7, bold=True)
        font_wh_bold = openpyxl.styles.Font(name='Arial', size=7, bold=True, color='ffffff')

        # Alignmentの設定
        al_cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
        al_lb = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
        al_c7 = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
        al_lt = Alignment(horizontal='left', vertical='top', wrap_text=True)
        al_ct = Alignment(horizontal='center', vertical='top', wrap_text=True)

        #----------------
        # header部 塗りつぶし
        #----------------
        for c in range(3, 12):
            for r in range(2, 6):
                self.tables_ws.cell(row=r, column=c).fill = fill_black
                self.tables_ws.cell(row=r, column=c).font = font_wh_bold

        #----------------
        # ルール部 塗りつぶし
        #----------------
        #条件部が記述してあるカラム毎に塗る
        for c in range(3, 3 + self.len_condition):
            for r in range(7, 19):

                # デフォルトの設定
                self.tables_ws.cell(row=r, column=c).font = font
                self.tables_ws.cell(row=r, column=c).border = border 

                # 各行を設定
                if r == 7:
                    self.tables_ws.cell(row=r, column=c).fill = fill_black
                    self.tables_ws.cell(row=r, column=c).font = font_wh_bold
                    self.tables_ws.cell(row=r, column=c).alignment = al_c7

                if r == 8:
                    self.tables_ws.cell(row=r, column=c).alignment = al_cc

                if r in (9, 10):
                    self.tables_ws.cell(row=r, column=c).alignment = al_lb

                if 8 <= r < 11:
                    self.tables_ws.cell(row=r, column=c).fill = fill_orange

                elif 12 <= r < 19:
                    self.tables_ws.cell(row=r, column=c).font = font_bold
                    self.tables_ws.cell(row=r, column=c).alignment = al_lt
                    # 条件部に水色を塗る
                    if c < 3 + self.len_condition:
                        self.tables_ws.cell(row=r, column=c).fill = fill_lightblue
                        # フォーマットの設定は最後に行う
                        self.tables_ws.cell(row=r, column=c).number_format = openpyxl.styles.numbers.FORMAT_TEXT

        #----------------
        # アクション部・条件部 セル内位置
        #----------------
        
        #D列からS列まで上左右中央揃え
        action_start = 3 + self.len_condition
        action_end = self.len_condition + self.len_act + 10
        for c in range(action_start, action_end):
            for r in range(12, 19):
                    self.tables_ws.cell(row=r, column=c).alignment = al_ct
        
        #D列からF列まで上左揃え
        action_start = 3 + self.len_condition
        action_end = 3 + self.len_condition + 3
        for c in range(action_start, action_end):
            strs = self.tables_ws.cell(row=11, column=c).value.splitlines()
            if not self._get_action_count_info(strs):
                for r in range(12, 19):
                    self.tables_ws.cell(row=r, column=c).alignment = al_lt
        
        #H列・I列の上左揃え
        action_start = 3 + self.len_condition + 4
        action_end = self.len_condition + self.len_act - 9
        for c in range(action_start, action_end):
            strs = self.tables_ws.cell(row=11, column=c).value.splitlines()
            if not self._get_action_count_info(strs):
                for r in range(12, 19):
                    self.tables_ws.cell(row=r, column=c).alignment = al_lt


    def _set_width(self):
        """
        [メソッド概要]
        11行目の幅を自動調整する
        日本語3byte, 英語1byteを考慮していい感じの幅にする
        """
        for c in range(3, self.len_condition + self.len_act + 1):
            strs = self.tables_ws.cell(row=11, column=c).value.splitlines()

            if self._get_action_count_info(strs):
                self.tables_ws.column_dimensions[get_column_letter(c)].width = 2.88
                theme = self.tables_ws.cell(row=11, column=c).value
                # TODO 英語化対応で別途処理が必要
                self.tables_ws.cell(row=11, column=c).value = theme[5:]
                self.tables_ws.cell(row=11, column=c).alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

            else:
                maxlen = 0
                for s in strs:
                    l = len(s) + (len(s.encode('utf-8')) -len(s)) / 2
                    maxlen = l if maxlen < l else maxlen

                self.tables_ws.column_dimensions[get_column_letter(c)].width = maxlen

    def _get_action_count_info(self, strs):
        """
        [メソッド概要]
        アクションリトライ間隔、アクションリトライ回数、アクション抑止間隔、アクション抑止回数の判定処理
        """

        if len(strs) >= 2 and (strs[0] in get_message('MOSJA11069', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11070', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11071', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11072', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11073', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11074', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11076', self.lang, showMsgId=False)
                                or strs[0] in get_message('MOSJA11078', self.lang, showMsgId=False)):
            return True
        else:
            return False

    def _create_tables_sheet(self):
        """
        [メソッド概要]
        Tablesシートを作成する
        """
        self._create_condition()
        self._create_header()
        self._set_styles()
        self._set_width()

    def _protect(self):
        """
        [メソッド概要]
        Tablesシート全体の処理
        """
        # パスワードを取得して復号
        password = System.objects.get(config_id='PROTECTION_PW').value
        cipher = AESCipher(settings.AES_KEY)
        password = cipher.decrypt(password)

        # パスワードをセット
        self.tables_ws.protection.password = password
        self.example_ws.protection.password = password
        self.lists_ws.protection.password = password
        # シートを保護する
        self.tables_ws.protection.enable()
        self.example_ws.protection.enable()
        self.lists_ws.protection.enable()

        # 特定のセルの保護を解除
        self.tables_ws.protection.formatColumns = False
        self.tables_ws.protection.formatRows = False
        unprotection = Protection(locked=False, hidden=False)
        for c in range(1, 3 + self.len_condition+self.len_act):
            for r in range(12, 19):
                self.tables_ws.cell(row=r, column=c).protection = unprotection

        # 折り畳む
        self.tables_ws.row_dimensions.group(2, 10, hidden=True)

    def create_decision_table(self):
        """
        [メソッド概要]
        ディシジョンテーブルを作成する
        [戻り値]
        bool : 実行結果
        """
        try:
            self._create_tables_sheet()
            self._protect()
            self.wb.save(self.save_path + self.table_name + '.xlsx')
        except Exception as e:
            logger.system_log('LOSM00001', traceback.format_exc())
            return False
        else:
            return True


class DecisionTableCustomizer:

    def __init__(self, input_filepath):
        self.wb = openpyxl.load_workbook(input_filepath)
        self.lang = _get_system_lang_mode()

    def custom_action_type(self):
        """
        [メソッド概要]
        アクション部のバリデーション
        アクション種別のセルに使用可能なアクション種別のリストを作成する。
        <builtin_formats 参照>
        https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
        """
        tables_ws = self.wb['Tables']
        header_row = tables_ws[11]
        target_header_text = get_message('MOSJA11066', self.lang, showMsgId=False)

        action_type_col = 0
        for cell in header_row:
            if cell.col_idx == 1:
                # A列は無視
                continue
            if cell.value == target_header_text:
                # hitしたら終了
                action_type_col = cell.col_idx
                break
            if cell.value is None or cell.value == '':
                # 空白を終端とする
                break

        if action_type_col == 0:
            # hitできずに来たらエラー
            raise Exception('No match action type col.')

        dti = ActionType.objects.filter(disuse_flag='0').values_list('driver_type_id', flat=True)
        rs  = DriverType.objects.filter(driver_type_id__in=dti).values('name', 'driver_major_version')
        ati = [get_message('MOSJA03149', self.lang, showMsgId=False)]
        for r in rs:
            name_version = r['name'] + '(ver' + str(r['driver_major_version']) + ')'
            ati.append(name_version)

        acts = ','.join(ati)
        acts_type_dv = DataValidation(type="list",formula1='"{}"'.format(acts))
        for r in range(12, 19):
            acts_type_dv.add(tables_ws.cell(row=r, column=action_type_col))

        for data_validation in tables_ws.data_validations.dataValidation:
            # TODO 1シート内に入力規則が複数来たらバグる
            tables_ws.data_validations.dataValidation.remove(data_validation)

        tables_ws.add_data_validation(acts_type_dv)

    def output(self):
        try:
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()

            return stream
        except Exception as e:
            # 呼び元で対応してね
            logger.system_log('LOSM00001', traceback.format_exc())
            raise

