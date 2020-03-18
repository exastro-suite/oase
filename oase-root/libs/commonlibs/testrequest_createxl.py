# Copyright 2019 NEC Corporation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# 
#     http://www.apache.org/licenses/LICENSE-2.0
# 
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#



"""
[概要]
   テストリクエスト用のエクセルファイルを作成する

"""


import os
import sys
import openpyxl
import pprint
import django
import traceback
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from libs.commonlibs.define import *
from libs.commonlibs.dt_component import DecisionTableComponent
from libs.webcommonlibs.sys_config import System
from web_app.models.models import DataObject, RuleType, ConditionalExpression
from web_app.templatetags.common import get_message

from libs.commonlibs.oase_logger import OaseLogger
logger = OaseLogger.get_instance() # ロガー初期化

class TestRequestXlFactory:
    """
    [クラス概要]
    TestRequest用エクセルを作成するクラス
    rule_type_idからルール種別、データオブジェクトを取得し、
    template_testrequest.xlsxファイルを基に作成する。
    """
    def __init__(self, rule_type_id, table_name, save_path, request):
        logger.logic_log('LOSI00001', 'rule_type_id: %s, table_name: %s, save_path: %s' % (rule_type_id, table_name, save_path))

        if rule_type_id and  not isinstance(rule_type_id, int):
            raise TypeError('rule_type_id:expected int, got %s' % type(rule_type_id).__name__)

        try:
            rule_type = RuleType.objects.get(rule_type_id=rule_type_id)
            tmp_data_list = DataObject.objects.filter(rule_type_id=rule_type_id).order_by('data_object_id')
            checker = {}
            for dataObj in tmp_data_list:
                if dataObj.label in checker:
                    continue
                checker[dataObj.label] = dataObj
            self.data_object_list = checker.values();

        except Exception as e:
            logger.system_log('LOSM00100', rule_type_id, traceback.format_exc())

        # 初期化
        self.rule_type_id = rule_type_id
        self.table_name = table_name + '_testrequest'
        self.save_path = save_path
        self.len_condition = len(self.data_object_list) # 条件部の数
        self.request = request

        # テンプレート読み込み
        cur_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
        self.wb = openpyxl.load_workbook(cur_dir + '/template_testrequest.xlsx')
        self.tables_ws = self.wb['Values']
        self.example_ws = self.wb['記述例']
        request_row_max = System.objects.get(config_id='REQUEST_ROW_MAX').value
        self.tables_ws['B6'].value = '記述できるリクエスト数の上限は %s 件です' % request_row_max

        # エクセルの列の最小と最大
        self.col_min = 2 # B列
        self.col_max = 4 # D列

        # エクセルの行の最小と最大
        self.row_min = 2
        self.row_max = 4

        logger.logic_log('LOSI00002', 'self.rule_type_id: %s, self.table_name: %s, self.save_path: %s, self.len_condition: %s' % (self.rule_type_id, self.table_name, self.save_path, self.len_condition))

    def _create_condition(self):
        """
        [メソッド概要]
        条件部を作成する
        """
        logger.logic_log('LOSI00001', 'None')

        # E列から条件部分のカラムを追加する
        self.tables_ws.insert_cols(self.col_max + 1, amount=self.len_condition - 1)

        #labelの数だけ条件部を作成する
        # D列から条件部作成
        for i, d in enumerate(self.data_object_list):
            self.tables_ws.cell(row = self.row_min, column = i+self.col_max, value = d.conditional_name)

            try:
                # コメント追記文言取得
                mosja = ConditionalExpression.objects.get(conditional_expression_id=d.conditional_expression_id).example
                example = get_message(mosja, self.request['lang'], showMsgId=False)

            except ConditionalExpression.DoesNotExist:
                logger.system_log('LOSM00101', d.conditional_expression_id)

            except Exception as e:
                logger.system_log('LOSM00102', d.conditional_expression_id, traceback.format_exc())

            # コメント追記カラム情報取得
            cellno = DecisionTableComponent.convert_rowcol_to_cellno(self.row_min-1, i+self.col_max-1)

            # コメント追記
            self.tables_ws[cellno].comment = Comment(example, 'Comment Author')
            self.tables_ws[cellno].comment.text

        logger.logic_log('LOSI00002', 'None')

    def _set_styles(self):
        """
        [メソッド概要]
        セルの塗りつぶし
        """

        logger.logic_log('LOSI00001', 'None')

        #----------------
        # 初期設定
        #----------------
        # 色の設定
        fill_lightblue = PatternFill(fill_type='solid', fgColor='ccffff')

        # 枠線の設定
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

        # fontの設定
        font = openpyxl.styles.Font(name='Arial', size=7)
        font_bold = openpyxl.styles.Font(name='Arial', size=7, bold=True)
        font_wh_bold = openpyxl.styles.Font(name='Arial', size=7, bold=True, color='ffffff')

        # Alignmentの設定
        al_cc = Alignment(horizontal='center', vertical='center', wrap_text=True)

        #----------------
        # ルール部 塗りつぶし
        #----------------
        # 条件部が記述してあるカラム毎に塗る
        # D列から条件部分ループ
        for c in range(self.col_max, self.col_max + self.len_condition):
            # 2行目から4行目までループ
            for r in range(self.row_min, self.row_max + 1):

                # デフォルトの設定
                self.tables_ws.cell(row=r, column=c).font = font
                self.tables_ws.cell(row=r, column=c).border = border 

                # 2行目から4行目を設定
                if self.row_min <= r <= self.row_max:
                    self.tables_ws.cell(row=r, column=c).font = font_bold
                    self.tables_ws.cell(row=r, column=c).alignment = al_cc
                    # 条件部に水色を塗る
                    if c < self.col_max + self.len_condition:
                        self.tables_ws.cell(row=r, column=c).fill = fill_lightblue
                        # フォーマットの設定は最後に行う
                        self.tables_ws.cell(row=r, column=c).number_format = openpyxl.styles.numbers.FORMAT_TEXT

        logger.logic_log('LOSI00002', 'None')

    def _set_width(self):
        """
        [メソッド概要]
        2行目の幅を自動調整する
        日本語3bite, 英語1biteを考慮していい感じの幅にする
        """

        logger.logic_log('LOSI00001', 'None')

        # C2セルから条件部分ループ
        for c in range(self.col_min + 1, self.col_max + self.len_condition):
            strs = self.tables_ws.cell(row=self.row_min, column=c).value.splitlines()
            maxlen = 5 # 40ピクセルくらい

            for s in strs:
                l = len(s) + (len(s.encode('utf-8')) - len(s)) / 2
                maxlen = l if maxlen < l else maxlen

            self.tables_ws.column_dimensions[get_column_letter(c)].width = maxlen

        logger.logic_log('LOSI00002', 'None')

    def _create_values_sheet(self):
        """
        [メソッド概要]
        Valuesシートを作成する
        """

        logger.logic_log('LOSI00001', 'None')

        self._create_condition()
        self._set_styles()
        self._set_width()

        logger.logic_log('LOSI00002', 'None')

    def create_testrequest_table(self):
        """
        [メソッド概要]
        テストリクエスト用エクセルを作成する
        [戻り値]
        bool : 実行結果
        """
        logger.logic_log('LOSI00001', 'self.rule_type_id: %s, self.table_name: %s, self.save_path: %s, self.len_condition: %s' % (self.rule_type_id, self.table_name, self.save_path, self.len_condition))

        try:
            self._create_values_sheet()
            self.wb.save(self.save_path + self.table_name + '.xlsx')
        except Exception as e:
            logger.logic_log('LOSM00001', e)
            return False
        else:
            logger.logic_log('LOSI00002', 'True')
            return True

